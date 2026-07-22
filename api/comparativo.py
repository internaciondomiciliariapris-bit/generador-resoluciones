"""
api/comparativo.py  —  Vercel Python Serverless Function
Genera el CUADRO COMPARATIVO de audifonos (.xlsx) calcado al modelo PRIS.

Archivo NUEVO y aislado: no toca api/generar.py ni ningun otro endpoint.
Requiere en requirements.txt:  openpyxl

Recibe POST con JSON:
{
  "expte": "5658/410/J/2026",
  "paciente": "JUAREZ CLARA ROSA",
  "fecha_adj": "21/07/2026",
  "cantidad": 2,
  "cotizaciones": [
      {"nombre": "IAR Argentina", "precio": 1570000},
      {"nombre": "GAES. S.A", "precio": 1690600},
      {"nombre": "OPTICA GIORLENT (GRUPO VISTALLI S.R.L)", "precio": 1380000}
  ],
  "idx_ganadora": 2,
  "nro_convocados": 4,
  "firmas_presentaron": "IAR ARGENTINA/GAES S.A/OPTICA VISTALLI"
}
Devuelve el .xlsx como descarga.
"""
from http.server import BaseHTTPRequestHandler
import json, base64, io, re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ---- Logo del encabezado (Ministerio de Salud Publica / Gobierno de Tucuman) embebido ----
LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAW8AAAA1CAYAAAH5WVqnAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAAIdUAACHVAQSctJ0AAClPSURBVHhe7Z0JnBTFucB7ODyiJkbxjkcUr5gYzwTzNOQl0RAEWXZm+hjwQAEPPGK8QNHlENAoCrI709dyiSIQxBvFgyin4n2jgsJyiU+55Nhdln7f91VVT/VM7+7spct7/f/9arrr6Oqe6q+/rqqu+kr5QdDsqUrnkr24b/dg1KK1nrekLblR8ypqlKTzHY9qQXRrJt9rFHjRpw99zjttCHPoRxdLuR5PAnfDfYrvKYph36fodoK26GTEHdPdy/Pimptjrn/MQycQfh7dYAJ/mNCdLxQ1sx13MdJP0M36iZJypiiG8xaFJTK/o/CUO0ykCaTPgV8vxS9Ztt67Y9oS5jfMSp4k5GIkNKc/bVPOaCjld/20uv0RbVuSO+eu8GQHhfQaj2r9UEnFrT9w726E4XzJ91o/WMr3LFirPzypE6nCEQtWM/k0TJu2LQI+FLp9N/c1nKT93fCFq5ah6hOqUKhDnoJx0b370dawX1BU62xSBkXjDoTnYARTidaVsN9b0axneTqmBltMHaqZ3wrVlxzzkle9s8Y7YsAj5D/6uqkevElP4CmbSkkbxXA3xgxnB/roYVLtwbRvOFtxC3L6OwynuGTpvrRFNNehrUzCurWqqqoG1d/UBV/QMfc++S6pQ54iQEzLbPDzRj+dw+lNW9X52g/TrPGgGkeIdC0CZj7wxS+rA6oQ0a0PaNvaiRmZSpJZ3byOB/0/R3MySk/zMCVR2hEf+gP62t7WhXt6Ixau8TqNnJMVuYimIQpRVDSFO/62WV7nW26B18WaQDjUl2KkTQ1nBmWw26HbO0AvWODmx1LmTh7KMKyVfA+V5wpQsN/GVJuUmY/mDMcNPLLryY9gWtWkumldyIV9O69nIvimWPT5Om/Pi8vJj28MAaYdOa9iAqiHU+HtcTtlJOhcui+9UQzrIXCdKUyzbobr6cNeg1AHRWVfZB5Hfnw9Gs4WePeza8cbqTtb4IUxDY6fS2FhJMzfwPEnUB7wbj4NNPpUyGgjRXYavTe9jQB6GyWtn1G42KrpX8MbYTSdHKHKuvUwFOBO5Vcle1C6i1x4j5e0oQIyMsezdO4wfGspWvo88hfzcKTYPZrv1YkocF6W9LZCsMBxHx1W+rHAhR+pqamZBAVziKKZUygjQcr6K6bBLfl1+xXcUJjYanYPeJveJYcrRZOwTnIj+gNpOXI4EpZmt0Fc9EmDnvQKcec98PIFIAQ3Q8ttHWUQ0UhUtwJcnPuUtqnxk2Kp8n/v3+/hCylANy/cLSWq1aNbLt/Lgmogab7HfRHNTs/0wfTC1p38wo9oHoS60Ce96Q2ft4p6EPbvN5lVBenl7ZwJDe9bKXFE/bTvPXEY3w2iO29Btexj3MXCvfWFZX7vNbpeQ6+gQh82r+K3mCbS4wUSWuB6eqmi2iUlC1cegIUq3LDXKrzL777UO13qeRcOD9s9C92wTuJ7+WiuyfewXn4m38uiOc/zvSC6cz3fyyOswEXByQUq9+vJTsSPXLh2FrQZjlIMey1lImM4Gah7/4P2dXsUOYT6/KS+Pd+PvWWw1cyU0tP+BY9l4H8x7Pu5DxtFw/08sNHl51EohtOV7wXBbjFsMGDHaBwzdu6IpdLrYLtC0TLHUBpsnenW5bQvk3JvVrC1imlzyCtwzV6odBtz1MgFaz8aPHcFNWgQvAnIIVc+7L384Wqv87CnyY+cfccTtUt5MvMhbRPO+UrcZo0xpPi+o5ViEC7spxHokAZbiWda7eFah0Cr+I/UMYt0GXsQbQWqncYNNBpZQzJpl8H/TPJWZlsKa43kFXgi8yvcHHfDI7tE5y86UeDbK6tpiwUux6Oj49UM6fQfBNV5XCkuP5b7gtCdNzJ3gQSxDmLezPYlosvYn4qmPtHN6iBLCzsemva4xbTYb4DN+5Q9WEmUH6RoZUf66eSecNX9NW3Pv28f+I3RvkRegSedD7HbobKqeh0WLObXphdrQn+6ZiNt0QkJx30BHo5+ykdgWOfCUzmY/rdm/kWOp32Ucvz22HnCXlTlRMlOpE/04/10pUfDk0wd4xiGwTHVqeH7LI3u0ks+m6Zr6aHooQDNdfx9IJawN0EhnkP7PKPsgWw/ZthboZCPihku9fpTPBW4cwAVqGaXU5juTKKDdEgv0J1t0CrczH0+oTrcMHfhFgsRvwoMnr6EtnKBO698QoUsPqJWVlbupELTzfz3heF0h8IciLt0fQhUKeE8oObsZxQ9k6Aw/Maj2pfAjaHOL5GWrkd1+gTC+HXIfjkc91sloQWedKqUhHtOpxFzhoX1m4Q5Oq41/9HWQliBE/gy0q1/dRn73J7y56RcN/CFrzZg8qiwC6Rtanwx382nc0k7KshEpid6D7z60d7YcdU25Y6neEQ3x0B19E3ui2g2kpnhpB/x/YIvMN2ezmMimh3dXQ4NjPwxL7o9me9FRIQQLzv1+NtYIw1dm5Tt7d9/iu8/aMBUz//iFhHRKikyz6FXDb7rkw5V/ZGY4axq27vc6zLo+kCnp+xOv/5Orx2kgebORH4YPBT2bD+/v2CvQUREM1FrCyEM7JrDMQyJ9J/RW+J5bYRmlh32b371Uoc84X5l5kleYvwbeelHvbrC7zyACvClSo/MMdA2XsJDIiIaT0ECrtnnMW1dVsVD8oaKNZfj2UODxKqhc7LewIgWAz9UJDJHQKNkEfmT9iYaGofo1nzc0BA51f4fSLtN0Uw21YGPelLibhWFq+5ECPsItlWQ1+cQE4O0lewYawgmjanmV5DnSjjHAAjfgmFKPH2eksLjze2BDwsIDm3D0WSUR4Z9nUlM25vS0jmd/6KwOqhXwFWnKwmamp6H3uue+2xPFET8aNIeqhp7Xzoh4DBt21754cK15V3QueHYNZ0n5PDf2INlHcVD8tHNp6GMTuc+ieltFc15kHsYmo3/IQb59lFw5JpAddkbRIz5xqGAhsOmyAjoY400+Aq77nXzJvpwo1pzFL3M7zGmoYJI7tDJBOSJvbQS/ocdAXbf4EONH3kSOR+DWgTN+jvfy4KfDhHdfpW2qew8HxCINbQ17JfoM2HSrYqpaWix25dROHsYPoO9NlB/rYqh0OvOKopTrWW0RcQf18xiBbvldWsb3OhzKUyA3dqGVa0kM99DHv+iMCwcw6mO6c5yuAETKKwO6hNwFDASMk4bzdouwgZOfYNmJ8gOw+MPvkhd98jxN07zDrpyMvd53h+HPuUdfs2UvOPOun0WHYsOBJ493F3GHiTCyF8bmvN8LJHeHsOeESzfjmP3pPB46e9ReGN4T7BrqstY9tkWBFzkS58ARJ+gJOB+PEtTmSfgujkFFNFVJOBiagjSI32eL+BKSTvFMHfGtMx62O5SEvYtLDwLCrg4D943/40FAh64BoxvFJpzp5Jk3z0KRrX+Ssc19hMq/tm43Y32UWgxLxrjm0PC6kxxZ/Zvr8RB0OmczUtDBbyyqpp9aq0FTItjf+VvVsLhOGD8YCgPvg5j245KpvlS6Z+LY8kfhmrOUPTMK3lONV+meNq3RtO+QKSR6W7/ErTwt3RvdGcRDb1CkulLKe2F8BYX4JsTwxL8+5mfX4486PYLeedBVOsaOMdyeABP4SFZMH3c7M19DD//hNPRLxCsCiQzKk+CmrUraMKHFTVzJT619OQClF6awhRLWJtwHweN+wWbNC+lfcP6AJ5W/8nyB5QLUs4UCu/CNQg0xsgPjSb8OEr7CJ8WBQX0OyVlD8Z9+tYGVQqoppCGpHRiBikOQtedSX64bs2kfd2dDlrkfrgp7A2CHyx19zFFT18q/8e6qLeKYrg6Oxerjq3fWnUjl0MC49BhFUP4CxVwOU7mtWWb2Jd5w2JVFPyKXx+G053S8g+6AgoT1UWOOCd59PS3/n4YhjOE4vnXdqJo0oEsX+dG9FJ+WNZQVaFwLHvcppzRtJVJZu8vbXVpFgfAjrdHsq0ko3SMZv9TeOgGUyLnC/9AfE3h3DsRJw7mAg4PSJV/rKj/IT0e3J/C0U9xGXZ8fQIuo2XeoTzweLgOSicLOIT71yWupzYB7/nAYSyddEyCdd2F/ce6KKyR6VzLzucsRu/n327vwmWxVmp27fK27qgOuJ01u3hs7cydO7cdnVMId9fSQ8lfH7UIONyjkzBcdlRuMqo9KjcNpUtyLa7ZC/PiWH2eoLDcPBFZwOENTfv4JpbR7OkU3n0MtQloP+WMoDgk7l5A+Yt8IhpGQQLOoZuYMLtwr/Kz/o9k67JNdIddPz07fFOzBsKb7Hvui4hoPA0RcB/VfgaFErRmtah3lnz44R6H3jB9++HXT/MKccfeOmttcrrH66yd21FjEPNU7RIWFhHRWihyThaamIRetUp5TD6a+byoolH6iIhWD3bLoeGplLuOCbm7FaoZdyiJ0XuDZh4qhNl3WNVhXXfR7LyI3RTVHAyN03cUnIKEs2qSITNzIiJ2J7Ar9O4Fq+lr5MA5y709eme84fOZ/x/PLsVutd3HulnE/1OwOoJDAZLuNzxEaW+UbxWf2S946DVvVFkXGlyllfT3ek95m8LJ+l3C8Y+hQe+GuZp7IyJ+VGKxhEvfCKh+DX4KTYw9qP3F4732vUzv3WePzBs9iG7GlDOp3o1DZOkYAmc9m5WUH7fJFxHRZPboNSFvLESdJOxfioahkrJP46HK8HkVXwqN3b5XJlSw0VUtbuft15dZHUE3fMHqhTwLrKP39fMOG6YQEdEQGiTc+rizFJwvV5SdzD58/pp/CkEVbti8Vd6evcu8XW8EBbsG/G1BcHPTo7vztdV8hGOyLX3xNdJLoyGxEU2iYOHmn5OpfswZuWjtW2GCKlyx+zpVP2bP+LW312UTvMumvheaTrih81f7feS+Bo+IaCyFCrc/GIwbpBk2v8INE9CmumGL1rCxGF3K2VDYAkzERjSFXDNNhtUBXs/MNFTSuZS2uehmiXLB6AO4L4vmsjHf9dHTPEOJu3/hvsJI2H4PRKEUJNwXpU/EwWPwvx9G78jXV50QJpj9ZnyQZ3arLnfOqDl5eUDdfVf/N99sT+c1Mh/Gkvim6MwGV9WFlhkO1zdH0Z0xzOwvR7U7wb2aDOGuUpzOGigS5rfQJcxsGWjuzbSV49GhNSocry/8Ymw9otPQgax9Ic26ge+RDQqIHwfpn1BSfE2PXOTziMFbiByOriWIJVwceYd9szGaZ6jZt1EXGKI5Gm1V+xlIsw1b/eTX7aEKzgrXM4NYuLOVGmP4dY5m1kw4BgcL8bht1GOgWkXKuemfK4a7AwriTMjzb5ATnNPexNIxO0MBVKeC4nR7hT/aUbceY2HpPPtCuRQk3IazRa4iDFu4cicKIoblzrjZ85LxtMVhsblx6MTsHpFOdsfcPJMEfOSiFdX8VGzUnG7TSM9aEZNHckna+esOxZmh8RjOuBKoEzopfx13OO3zxWPk/+uDAirDq2hQVfua4lSLWQDWTDYxQnWeVIofCtoADBtNKBQlgqMF+X0LvYbmhoQb0d3xcCFs4HmucBvWNbRFy2G9Mr+CdFy406dD3NsgyKT1YnFppodGU9dWkNOtT0m4xeB5IdxoelPGcJi9O0R3aUkNAQm3YZ5Ds0lEvkbO8TkUItxYyKKgSz709hBaFsNyeeqtFd72qp3eaYNmkr+f/Zo34vG3PeOhl8k/6LE3vO93VHmz360gvwBn6gjhRkcnBmJ6PUN5sVeFjz+nWUtY/jidDwgbliryIuFGW37osNx1h81okoVbxJPNP4wDAUbraoY7EtL5tg5JuJG4eRbIw5tZ4bY2Kn+9J2j2IuztKgt3EhrUOO0QCFxDCpRei5Bw/cmzPmJ4ZtJmkyR0+yraopV7rexk8JdA/fSnIIBvKGq6Qkmwpa/g9Xkz3ICNqMWVpHkbDWBHi249yo5U4pkeNG0JUbFaYl5A+zilTAPhT/I8ZPSyFTHdXquoOM2JrxujuU/FUEPR6L26KVS44Xoex/19+jysol+43Klks5Z86W3ZXuUL965duygdbhEU7o1bd3jPvL0icBzaWJTzPeT6mcykoM7H2ePrvTZ6PLg/Tj2Dh3tpTIf/bDiP8hi4d9ammGFtw355IfSIfC6atnYmF0JJuANpcNWEoObGuZpUvr5wC4x0dm6l7sIDAQ9o0vkGtsE5lxz/HDjXU80aEhLhfnxEwyioWoI3luxiKsrpg2ZemWvgU3YHXzkloLmRjyq+43tZzX3IVVNCjxfusH4Pn8HOXXoonf//HCVt6M2ScrNzM1sEnLOYeOiX3FcYeExj5zrialiq/Qn3sbySdj/uC0JxJjRaoU6O+6pzJY9pFuoVbm6XFK6XzPRu2F7Zh8tpKHK1pEP/hwOa52dXTKq1WpLLhspKNuPdcL6g89e28ChO0BZzD3Md9ser5k20L5OwExR24QNBC7iqXU514oRdBRr/Yh4K2pfnJ4N+MX8TIT+0u3KhY0MmcdPkbnC55OaLJM3e7PxJ5zt4PVXDq/92EIZ/U6VffqWhGQHZDjNAhadnzqcClE2D4b7sP3/yPvAKozl+4oZRuASF4Ux5NCLLpyvBq+lpugbMC4eB4jxB1IQ8b0pj2O+yRiWZOPZb3ux4qlPGaIEf2cwB1s8x/YU59TFt/JG5/7E2Cq5z84bysu8qT+XyR2D1guLBzftkTYOEe9O2HX7cuBc+5Dky8LTs3JmdGI/7oeDUP1xkCB2+/jGt8GMZG9yorwwIPIVxc9JQZfiO/Jo1icoSJ3NjBwJfvAjjcvOgMGEyArWwSCOsGyRLzxVhUG3JKjIEqj6otSm9ZmcNywP+MRhX5JxMgdC2Ij/Wq2iHmVcIZop1X828GgSPLe+azBgYTPso3Ib7AMuEgfvC71+M7l4ET+LncpwMhaXcrK0K0AIUptndAumhnib8uGVpzDIlzufh8YYn7aNwn2m1Z2ls+lxN18Me4m6xJDRuu5ayVaxonDTNumb/UbcHUHgtFCTcYtwHh8sf8cSbX/k3Y+5Hqxsk3Fj3FnGjn32f50iWnv2GIMWLHqj6EPdIpj7hLhp3Au3zbs4w6Bpy8qCwHOGGe0Fzev3rAGXE0knCXcxX/Tr/voPh/GNpXywzDFB6OtZ6irY6vDF94Za1tGbfQoG40Idus8m3uMSC7jzN9k1aboFlUpdw84tP2FfwqNA/jFA4CR40ErE3RLUfoIh6hJt6VAS69WkgLle4NYtN4EVTEDKqdSvLy54NmudJtt/03hIa/afbL8F/GYre/9ladS2XQ+/jVd+R+XF0y7/e5H2yaoNXtbPGKwVNHCbcL7xX4VVW7/SWrt1IjUhx7PxP1/IcCfa2JLPk1MDKs1EfSphw69Y9rBysDjwE0rnMZj6ZbJ+8j7g+HsuQJ32rfJCYBMtTNNizwk1eDE+UdsRdylsIN7xlyB/iMCkmkfyKErdPCabhF8IEjP9ZnE2eNJlAY18xVBtoP1e4/3DvfrnHUhyiu3NpPyxOgsJkzS24yA3mzfPAKBGOefs3SHWLRFyY5oYC/JD8In08sx1u1ossDP6jymb2N4twA5gX5cfhQlgnW0F4c2fE1wdobb/nIfec9SKVaYBkxjcy5OeZdMgKGAOrg9npb34afLh4fG4c3S88Tor3hVuC0grhNtI72XFB6Fi+eKfInyI4Md2su2oWEU6hwq2o48/mN+Fb9JaUeG1OGvj4rrD1DBrrTrnzaf+jk//wYzdfRERjKFi4kaQzkOr5IHssQFHa9LLoLdhUt9fFk6S3HdeE8RAzdxERhdIg4UZEgwitfHH++fznN4ctlFKou3b2UrZIbc/0wULYA2bNIiIaQ4OFG+lcuq8vhEn3aR6qnHv/3OvCbJXU5ordt/7ID8UG9mf8ocEeiMIakBERddE2Vd74V79m3eALue4eLo9oO7Fk1jE/uXzCA7He7gyIm9mu9/hp+10+edRpg57Lmu3F0Xvn38d6K7ChJVm2iohoNcT4GBASdDIDXItlXFz9SnXpowlLm6YGakRE6wfXajWcxUJ4SYCxK8ywipTk9LYKDigz3Nn+6MeIiFZP0jZAgFfisFvWZ++uYMNRobqC4ZKwM4E3v4e45YpRHjS2HxHR6sBP/obzsviKWTdQZcGZLPhBSi//PQ+MiIiIKBgcc361Ypirf9pvsnf5tPe92+YsJ+tfRc5ieoPe9vxyb8hrFd4tzy/zLiyd7+1hjK+BN+0HSnGm3vWWIiIiIiIaAo7m1aykorpswlIuup1oZ4zfnCh/3Z+5J9w/n/3M63D1o16Pwdd4a//zM+/oK+/xflvyrDcUFLicbsSCNTSPW0E7pqHf1LyYooKC15ziqH8gIiIiIowzrfY0gFqMexXOSO9ic7UkNPvVY299nJTxyIVrvDte/so7ZMBUMo3U444B3rpX9w+1B/bFiwd5p1w7lAze/eauZ7whr64kJY62eve+bCKOexzJz8DQrIlixSjhoJa/UTEyXXmKiIiIiNZFG2281iY1oTArS00had+r4EquvrLO4FqBpTRnUaJkrtfu7sXreoyYt/ZjtG6ESrdreoF3WN/7vY9mHx5Q1IW6Z6ad6rXtXe5dMf19UuIj51XUDF+wZs7wBV//gZ82C35+1u1Xg4oc51zaNOUtIiIiolWAoyobtTJcofRMHww17fd9RYgz42VzccCw/1QcDzXrd0Q3R5jrWjbfO/bKkaHKuS638MnjvJ9cUubdNPvz0HzRDVtYsXPEgoonSuZ+GbRQoI87PKDEdXtqYPpeRERExI9FiyjvrhMOVdCAllB8aKyqR2mebaK756+eOApqwWEKNcxh33X3zEJv/77jvev/pXvvP3ckdZPIyrpycXtv7uMneV0G3eB1uOZR74pprLZdqBu5cGX1iAWr8o0uqFY8qMjdx3hMRERExA9PMyvvNopmTwwoOdWaIddWS+Z+vS8ox+dEl0hT3bDXVtFSAacMftrrO+MDT6xF11Q3HGrjw+etYkahBT3RYov9if/fcFaukWFmZCMiIiJ+SJpReccUtLcmFDdZ+DHP4XHEqIXrbgxTlK3bVWy5e/G6TvwvMHTzJmHDg5S4Wv4kj4mI2M0wnK6sJmJWkl3DXNAQuZquYIJuZu11J0YfkPvBql6wv5SMcXoFzTYnw/mGcwf3Nh0cLYH/Jzm95YaUpdybcXYQ97UozaS8ocad+ROu6Unm3tBoKm4lRixa/a9w5Rjuhrxa4Q148lNa1WLqwi+85es3N9l9UPGd9/O+k71T73rWG/ji8tDzhrmR81bWDJvzxfH8rzDINn3pKWTUNOF0VHqmw4c6FopuuSCnOxTDepDKDw3Doj18tA2p2Q646QGZR8OnaLtdd5cqWrobtACOp2M052R4uQwAGdqiGPaS4LIzTp+Y7myhYZEJ90RyhnUS5HMG5DeX5hHL901zHXwhcx88x+Zx9Azj4hW4oEWuw3NjKwufEbQLj/bd89Kkz4PrWIzGf5WezvmUr2qdTfbicTac5rxBYblo5pvwTEzjPkbcHgrn2UHlgybD0WBwCv4vlZnzb9JHqjW61rnRHMh3I+UhrjEODiseuGwPmoPRMrMCeaTcJJUD3gP5v8lutwGVt1ipBO16as5bdPM0sxiE4RslbrKZWbplBZQ3mhQUq5gANJVRp3WCskKqwkOhO1tpFACCq5vo9iukwEm4rWWQZ1+y1hxCqPJWnVLI4wXuY6B5Ri09D4TqCfoPqlXEBXWR0mNC1kIaro5i4ANAy//A9bh3w/9YSQpdprtzCE7FzBv+JkiA4OJKIKrTl4dw4NxQLlnl3RnK096qJKxrmZ+DlqFxhe5EtvwaQ7Mob825C8tKOCibf/AY4u6Fa0bJynDwXGZF+vZpS7jBxdpBI7yYFg3xJse8FFgjQLB09QZv5hvLvfdXfstDsqCR3s7DniZDvUcMeKReM+u4aMbZdzwRWBFGuOGvrmYLhwji5seB/x1vxAzEi8YdDvL0JbiXAzZxi8pOVozygQGXyDBlZ9i3x5Lp75XkmKPIXxvxdDdauEPzFzzpE0s6VTEcTWPYL5HTrNdx4T+Iqwalh6b0s89eLcobjl8Vg+cu18FLg60KJJQ32tXV3bN8B0oaXBzk+wNwnyo9mUVuDCflrVlsXQM0/Y/PhmHf4o+7z1Xemg0vOsdGY288JBw4V0xz3uG+UEh5yyvh5IJLipGVd27JhSvvGBrozi0HzX4fngdmyXu3QFbedVGf8kYr5WIpKgEtiQQ3FpelQmTlTYKW81ZV7V5QyJtpOSWAlLduB5aagjepA+ny1xUgJT2d5YfKG2s2KIgyucpbte6B/7EyoOCRziX7stWFsoakA2jW3+k689aoQOVtv5JV3v3bw/m2hbZocCRHE2kW5Q01Nl+J6U5gLbf9+kz+fayX+40fL7n6lHenO5/wDuzHDEsfcc0jtIYc1sSPGPAorS9XdP8cb+5Ha7wdoNgRVLyLPvvau/De5739+kygdPv1mUhpD4fj2/ZyyVC1/tDLtBJSGEJ5514rur0umTjnD7fMz46U0eEFLY9X18saZ9UGasKgEKEGicoIWjG54MpVhv0FPBsP8RA4tzkBynqropXh85PTCsUp7dagGK5KlYRauICUNyhFvO484BjNmQbx2TX1aqt5G45voT0UobzJIpAELpWAS4zp6eVKl/Ksuatc5S2jZe6HvHDlqC0B5Y0L9OiZzXDM66HmZ3uN/QWUwbtw7FJqGdVBrcobWxGaM4VW99LMFA/N1ryxpRkR8WPRZOVdZNNyD8KBsD/IY4jkuBf793jwpZriB1/0GuPOuetJyrdo9ByqOefWvLFGfdjVj1AaVMzfbtnOYxhY8+7Qf7LX4/4XQPmP9/5ryFOh5ynU9bx3NlvdC6HZodlhkPT/G9oNGPHDg11duKRjwv4zvQTyXn4REbsBTVbePcYfSd03hrmeWki6Gegy2bh1e/9t2ytrxDqeDXWzlrC1jORuk6++2UJK/aeXT6SaONao20AaXA26Hfj3gpr2GYMep6VicrtNctcVbairrKzMKm9ceE2zHoUa5nr6/9RNFk2vb7Ukpu0NLZH3eJfMWkUtDyxa1/xg53uIGTvhFL0M12y6XXQTNAY/P2yq47I7LYXKFkHzrx3fgHWRcqbI6fMcfeAwNyhqWVGteaFdSzWThodrF30k0e3reEyWlE0LcaGDm/qukpCadHWRcml9LjoOu4Cw3xwfaGjeZfOTVgpvhTRHt0kswVYRpP+LzWKJTdXVf6+sqtrCK8Kh1NTsIsVYVc0UpFjIGQnr8/50zUaqZYtzhjlct2vJsvV19nnjeXbWZM+L11EXmyurvnp/3fbsB0DsttIz2UU72IIeDaPYOklJmavla6/PKbSIs3U2Hc8XwqNw7E6sDb44XjYP3t8s6DxhLwWXccVFrqV0tTl45j7xv0MZzpBAXCJ9IoXnUjTpQD8d/gfduZHHACFLOmFXrJY5hifIEi/9PS3dJKfNXWkyF1zCN+EEuu+g7JYo3ceEdCFlkdPTMbhcKz7nufDF/Px0UCh9fE8CajcXjA58wacPHFhQuNS8jG4+DcdulTMjZ2R2wk1fBi8DWodMkI2XlDdfFJBcwtpEYQI1/Ws/DhwPZWjl/UGIVoCyDBQuhK3Fvq5AWAOUNyjKjXmz2xIufhHe6KfBj3i44KDh+v+dp2RgXx3UDv04sbpiUHmvhTxehReD9FCiTVhnC/yHqcpF7uF0DBKmvFl/3yQ/XKztJsA+cBA0ELaAbQ96IFOZdXB/sqtNqs4MeOi+l9ORw2V5cW3fuN2Np2w0zfPB0hovXx/eAx6Da8rFtlXtzC7hmYO8SqhwWIPGFUORllLe8gqjspNXG82luro6uM6y6pTKx8JLvzOPaRqG0x0rG36+hjWQx+TTVOWNXT+6/UIgLp4+j+LCQJ2DFSKZllLewtEotsyfIP31UOZ5a/qRq0t5y2WEo0eK8aNpZrMflrBqXfNcpMl11A+eNPvzZCHKm4CC0uCB18oXwonXyAn8hLr7MQ1dEuBwH92EtxMuR2yXgBsM/i5wwW/4x/BFLhE/r6Yp75gfxkaN3M3Ds6guLUnrH9eQmjfc7FjS2sRc+ns8h58PfuikDw64AH0TlTe+3HKFr/iBY+XWDygr1q/bEOWtOt39MKyh6W52HfHaSKV/TteI9x/vo2beCef+O5zrAymvOhfyrI9mUd74QU11Z8NDgqvO/hlC8pqi23bs+IbrwAC7i/LesKNyLf8rWXAkhGZfBvd7AbgReUqtsbSE8uYrOvtphfLWxp0AcuzXSKnC0FAaobxRCQdHJUnKG/WHWDQV+6alihQ8N1uVuFlMcSmpJR+mvPFDpiZV7nC5a3yehDPS/ixgKOO5pD9yEPGUJgX3GMERM0l7U/ZYHBSRfeYpTOk+uiOc8CMeiWMOs2NJ8UunZs/zE+PwPOx85ytvUVgxCIGAKZVF2fT1KG/NusNXkFiYyTQKUIwUFF/7WzhKD8Q0R7JhgUrQ/A2PQoXfSVaAlKapNe8wDKerXFuFcinzl+rGsZxSiyBUeaNi1dyJgRuJC/xLowqUnrzJ2hDl3fn+Dn4YlieOdpFtfMTNv0C5Zmgfm7CGs8pPn2sUSS7n1qC8kaJxB8Z0M1ub0TPrafUKzprNmw/auXNn/WMDC6AhyrupQG1717bqmgdgN/tCUu1O8P+CLcuQB7/RNER5oywamfV+WuwaRKuJicwRJEd6eQKeh5ehQlPtp8ERWzLdxhwV0zIb/HisnerOa6Q4e0HNHP8bVv7UzCNY+UFlSl0sovsAu2wNSUlCqx+e/8dAwZ0Gx7aD858BeU2F8G1+GhxxEvi4W4vyFmC3x9/Mw7iPUZfy1qDMxLNONXeoABlOnzyn2n4lFdPDNV/FcyD8/DFeKG8B6i/NouXscx1PAWA/qmrpcKKJ8OAugpO+AYXxElzQ7TRONBccrK46Y1g6cxEoFguOOxmaHthkL2XOzQ7J88PgmOKHTuKhDKzp4RAeHJKj2bPguP+mQsweU8pTZsEaiZ4ZRgoNXhhw7Rmlu9NRSdg9A8d1npsd1xqGbl7qp6Xx3w2o2WDTT3dnUhmo9pPw8vibkpwE5QI1dJEnlhOSKAXB5GH4oNAkHLiJuvMiOx5qIzgRIhd6Ufh53ah0vA5eLiDoCetiPzxk2XVKp1uDqHxQ+SbTi6G80nDeE3gKRs+xv4Byf5CuIekshvxceCh/BdfWz89fd27lqRtFsylvBMe2g5wEhFh1boIYX/Edc9Pjlx484NHKX/xjhtfa3SEDHt18dIlkrIrWBctWgNj/y9Q+Prix6OZZoPAeh+d1FjkdnptCwGWeNOcuSP8KyRU5ewk9/1h5KhR8VnXnHtIdmvsOuaS9GPK2eYWn9o982K2iO73hWOy65dcATreeAN2hQyHW/gyL/4v/PZG+iIfWTjJzmX9M3B7HQ/G5nOaHJzND6tUzNIkwAzqSH6Pb2edBhKFLZHrx0Hx6lv+WXlgibURES9OsylsQNy+QVwCNqfYmeulkifWd9t6AQS9+6bVGd/Nzy6u18W8HzcUa5Z1lpQ3/b1vexKyIiIiIH4oWUd4y2I2UMLvEdHsbKPEqhc3282tu101bc9BvBj/xVcdb/73ruFtneT+W63jTrJrThz77Or8sQQyu92pU1tA6+p76hyMiIiJaAy2uvJHuZacqqkO2bYSDpu2XimYFarf79n+kw359p5jH3DTT+yHcIQOmbmh3iTzigaPbV8VyRvkoOO05OQEndERERET8+PwgylumCL+9OKvkD3JQI69R1MxYMlSWw03vrt1nxIK1o0fNX7205D+rt931n5VeQ1zJS8t3DZ23euPIRavfGb5wTbidmaKxxylGznA5HA2RbN1j9CMiIiJ+HPBrPX6Ij1srFRxqR6NszGlsdBPak3CfhVpwP6VjAaOMCgGHjiXsB+Bc+AG4hkYXoFEzsmFhPQUK+xMIQzO2tX+Qi4iIiIioBRxiqFpzFMPd6A9XlR2uf6nbWxQ9Y9CIIexXzxJTcPywIc18lJxCsxNxGr8zidJGRERERLQgqNCxpo6TgLCWLo0ZDnMQv1XRnPehJj0ZlPlQWgUnIiIiIqIFKJp0ICjcN8OUMTmcbIHGn1LmezgKRLkguOCDD46PjzujFVyowLDXos2RWvLDGvwYflRERDOhKP8LRb6vYPMD2BIAAAAASUVORK5CYII="

# ---------- Estilos calcados del modelo ----------
GRIS_GANADORA = PatternFill("solid", fgColor="D0CECE")
GRIS_CONFORME = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin")
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

F_EXPTE  = Font(name="Calibri", size=11, bold=True)
F_HEAD   = Font(name="Calibri", size=11, bold=False)
F_HEAD_B = Font(name="Calibri", size=11, bold=True)
F_DATA_B = Font(name="Arial", size=8, bold=True, color="FF000000")
F_DATA   = Font(name="Arial", size=8, bold=False, color="FF000000")
F_WIN    = Font(name="Calibri", size=11, bold=True)
F_CONF   = Font(name="Calibri", size=11, bold=False)
F_FIRMA  = Font(name="Times New Roman", size=12, bold=True)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

FMT_PESOS     = '"$"\\ #,##0.00'
FMT_PESOS_WIN = '_-"$"\\ * #,##0.00_-;\\-"$"\\ * #,##0.00_-;_-"$"\\ * "-"??_-;_-@_-'


def build_xlsx_bytes(d):
    expte              = str(d["expte"]).strip()
    paciente           = str(d["paciente"]).strip().upper()
    fecha_adj          = str(d["fecha_adj"]).strip()
    cantidad           = int(d["cantidad"])
    cotizaciones       = d["cotizaciones"]              # lista de {nombre, precio}
    idx_ganadora       = int(d["idx_ganadora"])
    nro_convocados     = int(d.get("nro_convocados", len(cotizaciones)))
    firmas_presentaron = str(d.get("firmas_presentaron", "")).strip()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws.column_dimensions["A"].width = 15.14
    ws.column_dimensions["E"].width = 22.43

    # Encabezado con logo arriba a la derecha
    ws.merge_cells("A1:E2")
    ws.row_dimensions[2].height = 21
    img = XLImage(io.BytesIO(base64.b64decode(LOGO_B64)))
    ws.add_image(img, "E1")

    # EXPTE / paciente / fecha
    ws.merge_cells("A3:E5")
    c = ws["A3"]
    c.value = f"EXPTE : {expte} - Paciente {paciente}" + (" " * 60) + f"(Adjudicacion de Fecha {fecha_adj})"
    c.font = F_EXPTE
    c.alignment = CTR
    for row in ws["A3:E5"]:
        for cc in row:
            cc.border = BORDER_ALL

    # Encabezado de tabla
    for rng in ("A6:A7","B6:B7","C6:C7","D6:D7","E6:E7"):
        ws.merge_cells(rng)
    ws["A6"] = "Detalle";  ws["A6"].font = F_HEAD
    ws["B6"] = "Cantidad"; ws["B6"].font = F_HEAD
    prov_cols = ["C", "D", "E"]
    for i, cot in enumerate(cotizaciones[:3]):
        col = prov_cols[i]
        cell = ws[f"{col}6"]
        cell.value = cot["nombre"]
        cell.font = F_HEAD_B
        if i == idx_ganadora:
            cell.fill = GRIS_GANADORA
    for col in ("A","B","C","D","E"):
        for r in (6,7):
            ws[f"{col}{r}"].alignment = CTR
            ws[f"{col}{r}"].border = BORDER_ALL

    # Datos
    ws["A8"] = "AUDIFONO"; ws["A8"].font = F_DATA_B; ws["A8"].alignment = CTR
    ws["B8"] = cantidad;   ws["B8"].font = F_DATA_B; ws["B8"].alignment = CTR
    for i, cot in enumerate(cotizaciones[:3]):
        col = prov_cols[i]
        cell = ws[f"{col}8"]
        cell.value = float(cot["precio"])
        cell.alignment = CTR
        if i == idx_ganadora:
            cell.font = F_WIN
            cell.fill = GRIS_GANADORA
            cell.number_format = FMT_PESOS_WIN
            ws[f"{col}9"].fill = GRIS_GANADORA
        else:
            cell.font = F_DATA
            cell.number_format = FMT_PESOS
    for col in ("A","B","C","D","E"):
        ws[f"{col}8"].border = Border(left=thin, right=thin, top=thin)
        ws[f"{col}9"].border = Border(left=thin, right=thin, bottom=thin)

    # Texto de adjudicacion
    ws.merge_cells("A10:E13")
    ws.row_dimensions[10].height = 15
    gan_nombre = cotizaciones[idx_ganadora]["nombre"]
    c = ws["A10"]
    c.value = ("CONFORME A LO DETALLADO EN EL CUADRO COMPARATIVO , SE ADJUDICA LA "
               "COMPRA DE LO SOLICITADO EN EL EXPTE DE REFERENCIA A LA FIRMA :  " + gan_nombre)
    c.font = F_CONF
    c.alignment = LEFT
    for row in ws["A10:E13"]:
        for cc in row:
            cc.fill = GRIS_CONFORME

    # Constancia de convocatoria
    lineas = [
        f"Se deja constancia que, habiendose solicitado cotizacion a {nro_convocados} proveedores del rubro, unicamente",
        f"las firmas comerciales: {firmas_presentaron}; presentaron presupuestos dentro",
        "del plazo establecido. Los restantes proveedores convocados no remitieron cotizacion ni",
        "emiteron respuesta alguna al requerimiento efectuado a la fecha de adjudicacion.-",
    ]
    for k, txt in enumerate(lineas):
        cell = ws[f"A{14+k}"]
        cell.value = txt
        cell.font = F_CONF

    # Firma
    firma = [
        "Firmado digitalmente:",
        "C.P.N Mariela Agustina Castillo",
        "Gerente Administrativo",
        "Direccion Gral. Prog. Integrado de Salud",
        "SI.PRO.SA",
    ]
    for k, txt in enumerate(firma):
        cell = ws[f"A{19+k}"]
        cell.value = txt
        cell.font = F_FIRMA
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[19+k].height = 15.75

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    if ws.sheet_properties.pageSetUpPr is None:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def nombre_archivo(d):
    expte = str(d.get("expte", ""))
    seg = [s for s in re.split(r"[/\\-]", expte) if s]
    exp_corto = "-".join(seg[:2]) if len(seg) >= 2 else (seg[0] if seg else "SN")
    pac = re.sub(r"[^A-Za-z0-9]+", "_", str(d.get("paciente", "")).strip().upper()).strip("_")
    cant = int(d.get("cantidad", 1))
    etiqueta = "AUDIFONOS" if cant > 1 else "AUDIFONO"
    try:
        precio_gan = int(round(float(d["cotizaciones"][int(d["idx_ganadora"])]["precio"])))
    except Exception:
        precio_gan = ""
    return f"CUADRO_COMPARATIVO_{exp_corto}_{pac}_{etiqueta}_X{cant}__{precio_gan}.xlsx"


class handler(BaseHTTPRequestHandler):
    def _cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            data = json.loads(self.rfile.read(length).decode("utf-8"))
            xlsx = build_xlsx_bytes(data)
            fname = nombre_archivo(data)
            self.send_response(200)
            self._cors()
            self.send_header("Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
            self.end_headers()
            self.wfile.write(xlsx)
        except Exception as e:
            self.send_response(500)
            self._cors()
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode("utf-8"))
